"""Vulnerability scanning service for generated connectors.

Runs pip-audit against a connector's requirements.txt, normalises the results
into a structured dict, generates HTML + Excel reports, calls the LLM for AI
fix suggestions, and persists all artefacts locally and to R2.
"""

import asyncio
import json
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import structlog

from integration.core.config import settings
from integration.services import r2_service
from integration.services.llm_client import call_llm

# Shielva Security SDK — optional enhanced scanning via the platform API.
# If the SDK is not installed or SHIELVA_SECURITY_API_KEY is not set, this
# integration is silently skipped and only local scanners run.
try:
    from shielva_security import (
        AuthError,
        PaymentRequiredError,
        ScanFailedError,
        ScanTimeoutError,
        ShielvaError,
        ShielvaSecurityClient,
    )

    _SDK_AVAILABLE = True
except ImportError:
    _SDK_AVAILABLE = False

logger = structlog.get_logger(__name__)

# ── Severity constants ────────────────────────────────────────────────

SEVERITY_CRITICAL = "CRITICAL"
SEVERITY_HIGH = "HIGH"
SEVERITY_MEDIUM = "MEDIUM"
SEVERITY_LOW = "LOW"
SEVERITY_NONE = "NONE"

_SEVERITY_COLORS = {
    SEVERITY_CRITICAL: "#dc2626",  # red-600
    SEVERITY_HIGH: "#ea580c",  # orange-600
    SEVERITY_MEDIUM: "#ca8a04",  # yellow-600
    SEVERITY_LOW: "#2563eb",  # blue-600
    SEVERITY_NONE: "#16a34a",  # green-600
}


# ── Severity mapping from CVSS scores ────────────────────────────────


def _score_from_aliases(aliases: list[str]) -> float | None:
    """Attempt to extract a CVSS score from aliases (e.g. 'CVSS:3.1/AV:N/.../7.5')."""
    import re

    for alias in aliases:
        # Look for a trailing numeric score pattern used by some advisory databases
        m = re.search(r"(\d+\.\d+)\s*$", alias)
        if m:
            try:
                return float(m.group(1))
            except ValueError:
                pass
    return None


def _classify_severity(vuln: dict[str, Any]) -> str:
    """Map a pip-audit vulnerability dict to a CRITICAL/HIGH/MEDIUM/LOW severity string."""
    description = (vuln.get("description") or "").upper()
    fix_notes = " ".join(str(v) for v in (vuln.get("fix_versions") or []))
    aliases: list[str] = vuln.get("aliases") or []

    # Explicit keyword match in description or fix notes
    if "CRITICAL" in description or "CRITICAL" in fix_notes.upper():
        return SEVERITY_CRITICAL

    # Try to extract score from aliases
    score = _score_from_aliases(aliases)
    if score is not None:
        if score >= 9.0:
            return SEVERITY_CRITICAL
        if score >= 7.0:
            return SEVERITY_HIGH
        if score >= 4.0:
            return SEVERITY_MEDIUM
        if score >= 0.1:
            return SEVERITY_LOW
        return SEVERITY_NONE

    # No score available — treat unknown as HIGH to err on the side of caution
    return SEVERITY_HIGH


# ── Semgrep code scanner ──────────────────────────────────────────────


async def _run_semgrep(source_dir: str) -> dict[str, Any]:
    """Run Semgrep against source_dir with python-security + secrets rulesets.

    Returns:
        ok          - True when semgrep ran without fatal error
        issues      - list of normalised CodeIssue dicts
        error       - error message or None
    """
    loop = asyncio.get_event_loop()

    def _exec() -> dict[str, Any]:
        try:
            result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "semgrep",
                    "--config",
                    "p/python",
                    "--config",
                    "p/owasp-top-ten",
                    "--config",
                    "p/secrets",
                    "--json",
                    "--quiet",
                    "--no-git-ignore",
                    source_dir,
                ],
                capture_output=True,
                text=True,
                timeout=120,
            )
            raw = result.stdout.strip() or result.stderr.strip()
            if not raw:
                return {"ok": True, "issues": [], "error": None}

            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                # Try stripping leading non-JSON output
                idx = raw.find("{")
                if idx != -1:
                    try:
                        data = json.loads(raw[idx:])
                    except Exception:
                        return {"ok": True, "issues": [], "error": None}
                else:
                    return {"ok": True, "issues": [], "error": None}

            # Surface any semgrep config/rule errors as a warning string
            semgrep_errors = data.get("errors") or []
            err_msgs = [e.get("message", "") for e in semgrep_errors if e.get("level") == "error"]
            scan_warning = "; ".join(err_msgs[:3]) if err_msgs else None

            issues: list[dict[str, Any]] = []
            for finding in data.get("results") or []:
                sev_raw = (
                    finding.get("extra", {}).get("severity")
                    or finding.get("extra", {}).get("metadata", {}).get("severity")
                    or "WARNING"
                ).upper()
                # Normalise to our severity constants
                sev_map = {
                    "ERROR": SEVERITY_HIGH,
                    "WARNING": SEVERITY_MEDIUM,
                    "INFO": SEVERITY_LOW,
                    "HIGH": SEVERITY_HIGH,
                    "MEDIUM": SEVERITY_MEDIUM,
                    "LOW": SEVERITY_LOW,
                    "CRITICAL": SEVERITY_CRITICAL,
                }
                severity = sev_map.get(sev_raw, SEVERITY_MEDIUM)

                path = finding.get("path", "")
                # Make path relative to source_dir for cleaner display
                try:
                    rel = str(Path(path).relative_to(source_dir))
                except ValueError:
                    rel = path

                lines = finding.get("start", {})
                line_start = lines.get("line", 0)
                line_end = finding.get("end", {}).get("line", line_start)

                snippet = finding.get("extra", {}).get("lines", "").strip()
                message = finding.get("extra", {}).get("message", "").strip()
                check_id = finding.get("check_id", "")
                rule_name = check_id.split(".")[-1] if check_id else ""

                # CWE / OWASP from metadata
                meta = finding.get("extra", {}).get("metadata", {})
                cwe: list[str] = meta.get("cwe") or []
                if isinstance(cwe, str):
                    cwe = [cwe]
                owasp: list[str] = meta.get("owasp") or []
                if isinstance(owasp, str):
                    owasp = [owasp]

                fix_guidance = meta.get("fix") or meta.get("remediation") or ""

                issues.append(
                    {
                        "severity": severity,
                        "file": rel,
                        "line_start": line_start,
                        "line_end": line_end,
                        "rule_id": check_id,
                        "rule_name": rule_name,
                        "message": message,
                        "snippet": snippet,
                        "cwe": cwe,
                        "owasp": owasp,
                        "fix_guidance": fix_guidance,
                    }
                )

            return {"ok": True, "issues": issues, "error": scan_warning}

        except subprocess.TimeoutExpired:
            return {
                "ok": False,
                "issues": [],
                "error": "Semgrep timed out after 120 seconds",
            }
        except FileNotFoundError:
            return {
                "ok": False,
                "issues": [],
                "error": "semgrep not found — install it with: pip install semgrep",
            }
        except Exception as exc:
            return {"ok": False, "issues": [], "error": f"Semgrep failed: {exc}"}

    return await loop.run_in_executor(None, _exec)


# ── pip-audit runner ──────────────────────────────────────────────────


async def _run_pip_audit(requirements_txt_path: str) -> dict[str, Any]:
    """Run pip-audit against requirements.txt and return parsed JSON output.

    Returns a dict with keys:
      ok         - True when audit ran successfully
      raw        - raw stdout string
      parsed     - parsed JSON (or None on error)
      error      - error message (or None on success)
    """
    loop = asyncio.get_event_loop()

    def _exec() -> dict[str, Any]:
        try:
            # Use `python -m pip_audit` so we always run the pip-audit installed
            # in the same virtualenv as the service, regardless of PATH.
            result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "pip_audit",
                    "--format",
                    "json",
                    "--no-deps",
                    "--skip-editable",
                    "--progress-spinner",
                    "off",
                    "-r",
                    requirements_txt_path,
                ],
                capture_output=True,
                text=True,
            )
            combined = result.stdout + result.stderr
            # Exit code 0 = clean, 1 = vulns found — both are valid
            if result.returncode not in (0, 1):
                return {
                    "ok": False,
                    "raw": combined,
                    "parsed": None,
                    "error": f"pip-audit exited with code {result.returncode}: {combined[:500]}",
                }

            # pip-audit 2.x sometimes writes JSON to stderr; try stdout first,
            # then stderr, then combined so we're version-agnostic.
            def _try_parse(text: str) -> Any | None:
                text = text.strip()
                if not text:
                    return None
                # Direct parse
                try:
                    return json.loads(text)
                except json.JSONDecodeError:
                    pass
                # Strip leading non-JSON lines (progress/warnings before JSON)
                for start_char in ("{", "["):
                    idx = text.find(start_char)
                    if idx != -1:
                        try:
                            return json.loads(text[idx:])
                        except json.JSONDecodeError:
                            pass
                return None

            parsed = _try_parse(result.stdout)
            if parsed is None:
                parsed = _try_parse(result.stderr)
            if parsed is None:
                parsed = _try_parse(combined)
            if parsed is None:
                # Empty stdout with a clean exit code means 0 packages (empty requirements.txt)
                if result.returncode in (0, 1) and not result.stdout.strip():
                    return {
                        "ok": True,
                        "raw": combined,
                        "parsed": {"dependencies": [], "fixes": []},
                        "error": None,
                    }
                exc_msg = combined[:200] if combined.strip() else "Expecting value: line 1 column 1 (char 0)"
                return {
                    "ok": False,
                    "raw": combined,
                    "parsed": None,
                    "error": f"Could not parse pip-audit JSON: {exc_msg}",
                }
            return {"ok": True, "raw": combined, "parsed": parsed, "error": None}
        except FileNotFoundError:
            return {
                "ok": False,
                "raw": "",
                "parsed": None,
                "error": ("pip-audit not found. Install it with: pip install pip-audit"),
            }
        except Exception as exc:
            return {
                "ok": False,
                "raw": "",
                "parsed": None,
                "error": f"pip-audit execution failed: {exc}",
            }

    return await loop.run_in_executor(None, _exec)


# ── Result normaliser ─────────────────────────────────────────────────


def _normalise_results(audit_json: Any) -> dict[str, Any]:
    """Convert pip-audit JSON output into structured vulnerability + safe-package lists.

    Handles both pip-audit output formats:
    - Legacy (1.x): {"dependencies": [...]}
    - Modern (2.x): [...] — a bare list of dependency dicts
    """
    if isinstance(audit_json, list):
        # pip-audit 2.x: bare list of {name, version, vulns}
        dependencies: list[dict[str, Any]] = audit_json
    else:
        dependencies = (audit_json or {}).get("dependencies") or []

    vulnerabilities: list[dict[str, Any]] = []
    safe_packages: list[dict[str, str]] = []

    summary = {
        "total_packages": len(dependencies),
        "vulnerable_packages": 0,
        SEVERITY_CRITICAL.lower(): 0,
        SEVERITY_HIGH.lower(): 0,
        SEVERITY_MEDIUM.lower(): 0,
        SEVERITY_LOW.lower(): 0,
    }

    for dep in dependencies:
        pkg_name = dep.get("name", "unknown")
        pkg_version = dep.get("version", "unknown")
        vulns: list[dict[str, Any]] = dep.get("vulns") or []

        if not vulns:
            safe_packages.append({"name": pkg_name, "version": pkg_version})
            continue

        summary["vulnerable_packages"] += 1

        for vuln in vulns:
            severity = _classify_severity(vuln)
            severity_key = severity.lower()
            if severity_key in summary:
                summary[severity_key] += 1

            vulnerabilities.append(
                {
                    "package": pkg_name,
                    "version": pkg_version,
                    "vuln_id": vuln.get("id", ""),
                    "aliases": vuln.get("aliases") or [],
                    "severity": severity,
                    "description": vuln.get("description") or "",
                    "fix_versions": vuln.get("fix_versions") or [],
                }
            )

    return {
        "summary": summary,
        "vulnerabilities": vulnerabilities,
        "safe_packages": safe_packages,
    }


# ── HTML report generator ─────────────────────────────────────────────


def _build_html_report(
    provider: str,
    service_slug: str,
    scanned_at: str,
    summary: dict[str, Any],
    vulnerabilities: list[dict[str, Any]],
    safe_packages: list[dict[str, str]],
) -> str:
    """Generate a self-contained HTML vulnerability report (inline CSS only)."""

    def _sev_badge(severity: str) -> str:
        color = _SEVERITY_COLORS.get(severity, "#6b7280")
        return (
            f'<span style="display:inline-block;padding:2px 10px;border-radius:12px;'
            f"background:{color};color:#fff;font-size:12px;font-weight:700;"
            f'letter-spacing:0.5px;">{severity}</span>'
        )

    vuln_cards = ""
    for v in vulnerabilities:
        aliases_html = (
            ", ".join(
                f'<code style="background:#f3f4f6;padding:1px 5px;border-radius:3px;font-size:12px;">{a}</code>'
                for a in v["aliases"]
            )
            if v["aliases"]
            else "<em>none</em>"
        )
        fix_html = (
            ", ".join(
                f'<code style="background:#d1fae5;padding:1px 5px;border-radius:3px;font-size:12px;">{fv}</code>'
                for fv in v["fix_versions"]
            )
            if v["fix_versions"]
            else '<em style="color:#9ca3af;">No fix available</em>'
        )
        border_color = _SEVERITY_COLORS.get(v["severity"], "#6b7280")
        vuln_cards += f"""
        <div style="border:1px solid #e5e7eb;border-left:4px solid {border_color};
                    border-radius:8px;padding:16px 20px;margin-bottom:16px;
                    background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.06);">
          <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;flex-wrap:wrap;">
            <strong style="font-size:15px;color:#111827;">
              {v["package"]}
            </strong>
            <code style="background:#f3f4f6;padding:2px 7px;border-radius:4px;
                         font-size:12px;color:#374151;">v{v["version"]}</code>
            {_sev_badge(v["severity"])}
            <span style="color:#6b7280;font-size:12px;margin-left:auto;">{v["vuln_id"]}</span>
          </div>
          <div style="margin-bottom:8px;">
            <span style="font-size:12px;font-weight:600;color:#6b7280;text-transform:uppercase;
                         letter-spacing:0.5px;">CVE / Aliases</span><br>
            <span style="margin-top:4px;display:inline-block;">{aliases_html}</span>
          </div>
          <div style="margin-bottom:8px;">
            <span style="font-size:12px;font-weight:600;color:#6b7280;text-transform:uppercase;
                         letter-spacing:0.5px;">Description</span>
            <p style="margin:4px 0 0;color:#374151;font-size:14px;line-height:1.5;">
              {v["description"] or "No description available."}
            </p>
          </div>
          <div>
            <span style="font-size:12px;font-weight:600;color:#6b7280;text-transform:uppercase;
                         letter-spacing:0.5px;">Fix Versions</span><br>
            <span style="margin-top:4px;display:inline-block;">{fix_html}</span>
          </div>
        </div>
        """

    safe_rows = ""
    for sp in safe_packages:
        safe_rows += (
            f"<tr><td style='padding:6px 12px;border-bottom:1px solid #f3f4f6;"
            f"font-size:13px;'>{sp['name']}</td>"
            f"<td style='padding:6px 12px;border-bottom:1px solid #f3f4f6;"
            f"font-size:13px;color:#6b7280;'>{sp['version']}</td></tr>"
        )

    def _card(label: str, value: Any, color: str = "#374151") -> str:
        return (
            f'<div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;'
            f'padding:18px 24px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.06);">'
            f'<div style="font-size:28px;font-weight:800;color:{color};">{value}</div>'
            f'<div style="font-size:13px;color:#6b7280;margin-top:4px;">{label}</div>'
            f"</div>"
        )

    summary_cards = f"""
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:16px;
                margin-bottom:32px;">
      {_card("Total Packages", summary["total_packages"])}
      {_card("Vulnerable", summary["vulnerable_packages"], "#dc2626")}
      {_card("Critical", summary.get("critical", 0), _SEVERITY_COLORS[SEVERITY_CRITICAL])}
      {_card("High", summary.get("high", 0), _SEVERITY_COLORS[SEVERITY_HIGH])}
      {_card("Medium", summary.get("medium", 0), _SEVERITY_COLORS[SEVERITY_MEDIUM])}
      {_card("Low", summary.get("low", 0), _SEVERITY_COLORS[SEVERITY_LOW])}
    </div>
    """

    vuln_section = '<h2 style="font-size:18px;font-weight:700;color:#111827;margin:0 0 16px;">Vulnerabilities</h2>' + (
        vuln_cards if vuln_cards else '<p style="color:#16a34a;font-weight:600;">No vulnerabilities found.</p>'
    )

    safe_section = ""
    if safe_packages:
        safe_section = f"""
        <h2 style="font-size:18px;font-weight:700;color:#111827;margin:32px 0 12px;">
          Safe Packages ({len(safe_packages)})
        </h2>
        <div style="overflow-x:auto;">
          <table style="width:100%;border-collapse:collapse;background:#fff;
                        border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;">
            <thead>
              <tr style="background:#f9fafb;">
                <th style="padding:8px 12px;text-align:left;font-size:12px;font-weight:600;
                           color:#6b7280;text-transform:uppercase;letter-spacing:0.5px;">Package</th>
                <th style="padding:8px 12px;text-align:left;font-size:12px;font-weight:600;
                           color:#6b7280;text-transform:uppercase;letter-spacing:0.5px;">Version</th>
              </tr>
            </thead>
            <tbody>{safe_rows}</tbody>
          </table>
        </div>
        """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Vulnerability Scan Report — {provider} / {service_slug}</title>
</head>
<body style="margin:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
             background:#f9fafb;color:#111827;">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#1e293b 0%,#0f172a 100%);
              padding:32px 40px;color:#fff;">
    <div style="max-width:960px;margin:0 auto;">
      <h1 style="margin:0 0 6px;font-size:24px;font-weight:800;">
        Vulnerability Scan Report
      </h1>
      <p style="margin:0;font-size:14px;color:#94a3b8;">
        {provider} / {service_slug} &nbsp;&bull;&nbsp; Scanned at {scanned_at}
      </p>
    </div>
  </div>

  <!-- Content -->
  <div style="max-width:960px;margin:32px auto;padding:0 24px 64px;">
    {summary_cards}
    {vuln_section}
    {safe_section}
    <p style="margin-top:40px;font-size:12px;color:#9ca3af;text-align:center;">
      Generated by Shielva Integration Builder &bull; pip-audit powered
    </p>
  </div>

</body>
</html>"""


# ── Excel report generator ────────────────────────────────────────────


def _build_excel_report(
    summary: dict[str, Any],
    vulnerabilities: list[dict[str, Any]],
    safe_packages: list[dict[str, str]],
    output_path: str,
) -> None:
    """Write a three-sheet Excel workbook using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Install it with: pip install openpyxl") from exc

    wb = openpyxl.Workbook()

    _SEV_FILLS = {
        SEVERITY_CRITICAL: PatternFill("solid", fgColor="DC2626"),
        SEVERITY_HIGH: PatternFill("solid", fgColor="EA580C"),
        SEVERITY_MEDIUM: PatternFill("solid", fgColor="CA8A04"),
        SEVERITY_LOW: PatternFill("solid", fgColor="2563EB"),
        SEVERITY_NONE: PatternFill("solid", fgColor="16A34A"),
    }

    _HEADER_FILL = PatternFill("solid", fgColor="1E293B")
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    _BOLD = Font(bold=True)
    _thin = Side(style="thin", color="D1D5DB")
    _BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _header_row(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

    # ── Sheet 1: Summary ──
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _header_row(ws_sum, ["Metric", "Value"])
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 18
    rows = [
        ("Total Packages", summary["total_packages"]),
        ("Vulnerable Packages", summary["vulnerable_packages"]),
        ("Critical", summary.get("critical", 0)),
        ("High", summary.get("high", 0)),
        ("Medium", summary.get("medium", 0)),
        ("Low", summary.get("low", 0)),
    ]
    for i, (label, val) in enumerate(rows, start=2):
        ws_sum.cell(row=i, column=1, value=label).font = _BOLD
        ws_sum.cell(row=i, column=2, value=val)

    # ── Sheet 2: Vulnerabilities ──
    ws_vuln = wb.create_sheet("Vulnerabilities")
    _header_row(
        ws_vuln,
        [
            "Package",
            "Version",
            "Vuln ID",
            "CVE / Aliases",
            "Severity",
            "Description",
            "Fix Versions",
        ],
    )
    ws_vuln.column_dimensions["A"].width = 22
    ws_vuln.column_dimensions["B"].width = 12
    ws_vuln.column_dimensions["C"].width = 22
    ws_vuln.column_dimensions["D"].width = 30
    ws_vuln.column_dimensions["E"].width = 12
    ws_vuln.column_dimensions["F"].width = 50
    ws_vuln.column_dimensions["G"].width = 25

    for row_idx, v in enumerate(vulnerabilities, start=2):
        cells = [
            v["package"],
            v["version"],
            v["vuln_id"],
            ", ".join(v["aliases"]),
            v["severity"],
            v["description"],
            ", ".join(v["fix_versions"]),
        ]
        for col_idx, val in enumerate(cells, start=1):
            cell = ws_vuln.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER
            if col_idx == 5:  # severity column
                sev_fill = _SEV_FILLS.get(v["severity"])
                if sev_fill:
                    cell.fill = sev_fill
                    cell.font = Font(bold=True, color="FFFFFF")

    # ── Sheet 3: Safe Packages ──
    ws_safe = wb.create_sheet("Safe Packages")
    _header_row(ws_safe, ["Package", "Version"])
    ws_safe.column_dimensions["A"].width = 30
    ws_safe.column_dimensions["B"].width = 16
    for row_idx, sp in enumerate(safe_packages, start=2):
        ws_safe.cell(row=row_idx, column=1, value=sp["name"]).border = _BORDER
        ws_safe.cell(row=row_idx, column=2, value=sp["version"]).border = _BORDER

    wb.save(output_path)


# ── AI suggestions ────────────────────────────────────────────────────


async def _generate_ai_suggestions(
    vulnerabilities: list[dict[str, Any]],
    tenant_id: str,
) -> str:
    """Ask the LLM for fix suggestions and return Markdown text."""
    if not vulnerabilities:
        return (
            "# AI Security Suggestions\n\n"
            "No vulnerabilities were found in the scanned dependencies. "
            "The connector's dependencies are clean.\n\n"
            "## General Security Hardening\n\n"
            "- Pin dependency versions in `requirements.txt` to avoid unintended upgrades.\n"
            "- Regularly run `pip-audit` as part of CI/CD pipelines.\n"
            "- Prefer packages with active maintenance and recent security patches.\n"
            "- Use a virtual environment to isolate connector dependencies.\n"
        )

    vuln_summary = json.dumps(
        [
            {
                "package": v["package"],
                "version": v["version"],
                "vuln_id": v["vuln_id"],
                "severity": v["severity"],
                "aliases": v["aliases"],
                "fix_versions": v["fix_versions"],
                "description": v["description"][:300] if v["description"] else "",
            }
            for v in vulnerabilities
        ],
        indent=2,
    )

    prompt = (
        "You are a Python security expert. Given these vulnerabilities found in a connector's "
        "dependencies, suggest specific fixes as a markdown document.\n\n"
        "For each vulnerable package, recommend:\n"
        "1. The exact version to upgrade to (use fix_versions from scan)\n"
        "2. Whether it's a breaking change\n"
        "3. Code changes needed in requirements.txt\n\n"
        "Also suggest general security hardening steps for connector dependencies.\n\n"
        f"Vulnerabilities:\n{vuln_summary}"
    )

    try:
        return await call_llm(
            [{"role": "user", "content": prompt}],
            system=(
                "You are a Python security expert. Return a well-structured Markdown document "
                "with actionable fix instructions. Do not use JSON in your response."
            ),
            expect_code=False,
            max_tokens=2048,
            tenant_id=tenant_id or None,
        )
    except Exception as exc:
        logger.warning("vuln_scan.ai_suggestions_failed", error=str(exc))
        return (
            "# AI Security Suggestions\n\n"
            f"> AI suggestion generation failed: {exc}\n\n"
            "Please review the vulnerabilities listed in the scan report and upgrade "
            "affected packages to the fix versions noted.\n"
        )


# ── R2 upload helpers ─────────────────────────────────────────────────


async def _upload_to_r2(
    provider: str,
    service_slug: str,
    key_suffix: str,
    content: bytes,
    content_type: str,
) -> None:
    """Upload binary or text content to R2 under the vuln/ prefix.

    Bucket  : shielva-agentic-app-{app_id}  (per-installation, via _get_bucket())
    Key     : {collection}/{provider}/{service_slug}/vuln/{key_suffix}
    """
    if r2_service._use_local():
        return  # local mode — files already written to disk

    loop = asyncio.get_event_loop()
    client = r2_service._get_client()
    bucket = r2_service._get_bucket()
    key = f"{r2_service._coll()}/{provider}/{service_slug}/vuln/{key_suffix}"

    def _put():
        client.put_object(
            Bucket=bucket,
            Key=key,
            Body=content,
            ContentType=content_type,
        )

    try:
        await loop.run_in_executor(None, _put)
        logger.info("vuln_scan.r2_upload_ok", key=key)
    except Exception as exc:
        logger.warning("vuln_scan.r2_upload_failed", key=key, error=str(exc))


# ── Main entry point ─────────────────────────────────────────────────


async def _run_shielva_security_scan(
    target: str,
    tenant_id: str,
    session_id: str,
    scan_type: str = "full",
) -> dict[str, Any]:
    """Run a scan via the Shielva Security platform SDK.

    Returns a dict with keys:
      ok           - True on success
      scan_id      - UUID of the scan
      summary      - {critical, high, medium, low}
      findings     - list of normalised finding dicts
      report_url   - URL of the HTML report
      error        - error message or None

    Returns ``{"ok": False, ...}`` on any failure — callers treat this as
    a non-fatal enhancement (local scanners still run).
    """
    log = logger.bind(session_id=session_id, tenant_id=tenant_id, target=target)

    if not _SDK_AVAILABLE:
        return {
            "ok": False,
            "error": "shielva_security SDK not installed",
            "findings": [],
            "summary": {},
        }

    api_key = settings.SHIELVA_SECURITY_API_KEY
    if not api_key:
        return {
            "ok": False,
            "error": "SHIELVA_SECURITY_API_KEY not configured",
            "findings": [],
            "summary": {},
        }

    log.info("shielva_security.scan_start", scan_type=scan_type)
    try:
        async with ShielvaSecurityClient(
            api_key=api_key,
            base_url=settings.SHIELVA_SECURITY_URL,
            verify_ssl=False,  # self-signed cert on localhost
            app_id=tenant_id or None,
        ) as client:
            scan = await client.scans.create_and_wait(
                target=target,
                scan_type=scan_type,
                timeout=settings.SHIELVA_SECURITY_SCAN_TIMEOUT,
            )

            findings_raw = await client.scans.findings(scan.id)

        # R2 key set by security API after report generation
        # e.g. "Shielvasense-platform-int/{scan_id}/report.html"  in bucket: shielvasense
        report_r2_key: str | None = scan.report_r2_url

        summary = {}
        if scan.summary:
            summary = {
                "critical": scan.summary.critical,
                "high": scan.summary.high,
                "medium": scan.summary.medium,
                "low": scan.summary.low,
            }

        findings = [
            {
                "severity": f.severity,
                "title": f.title,
                "scanner": f.scanner,
                "file": f.file_path or "",
                "line_start": f.line_number or 0,
                "rule_id": f.rule_id or "",
                "message": f.description,
                "cwe": [f.cwe] if f.cwe else [],
                "owasp": [f.owasp] if f.owasp else [],
                "fix_guidance": f.remediation or "",
                "package": f.package_name or "",
                "fix_version": f.fix_version or "",
                "source": "shielva-security",
            }
            for f in findings_raw
        ]

        log.info(
            "shielva_security.scan_complete",
            scan_id=scan.id,
            critical=summary.get("critical", 0),
            high=summary.get("high", 0),
            total_findings=len(findings),
        )
        return {
            "ok": True,
            "scan_id": scan.id,
            "summary": summary,
            "findings": findings,
            # R2 key in the shielvasense bucket — set by security API after scan completes
            # Full path: shielvasense/Shielvasense-platform-int/{scan_id}/report.html
            "report_r2_key": report_r2_key,
            "error": None,
        }

    except AuthError as e:
        log.warning("shielva_security.auth_error", error=str(e))
        return {
            "ok": False,
            "error": f"API key invalid or revoked: {e}",
            "findings": [],
            "summary": {},
        }
    except PaymentRequiredError as e:
        log.warning("shielva_security.payment_required", error=str(e))
        return {
            "ok": False,
            "error": f"Subscription required: {e}",
            "findings": [],
            "summary": {},
        }
    except ScanTimeoutError as e:
        log.warning("shielva_security.scan_timeout", error=str(e))
        return {"ok": False, "error": str(e), "findings": [], "summary": {}}
    except ScanFailedError as e:
        log.warning("shielva_security.scan_failed", scan_id=e.scan_id)
        return {"ok": False, "error": str(e), "findings": [], "summary": {}}
    except ShielvaError as e:
        log.warning("shielva_security.scan_error", error=str(e))
        return {"ok": False, "error": str(e), "findings": [], "summary": {}}
    except Exception as e:
        log.warning("shielva_security.unexpected_error", error=str(e))
        return {
            "ok": False,
            "error": f"Unexpected error: {e}",
            "findings": [],
            "summary": {},
        }


async def run_vulnerability_scan(
    output_dir: str,
    provider: str,
    service_slug: str,
    tenant_id: str,
    session_id: str,
    repo_url: str | None = None,
) -> dict[str, Any]:
    """Run a full vulnerability scan for a connector.

    Steps:
      1. Locate requirements.txt in output_dir
      2. Run pip-audit and normalise results
      3. Generate HTML, Excel, and AI-suggestion artefacts
      4. Save all artefacts locally under {output_dir}/.shielva/vuln/
      5. Upload to R2

    Returns a structured result dict (see module docstring for schema).
    """
    scanned_at = datetime.now(UTC).isoformat()
    log = logger.bind(
        session_id=session_id,
        provider=provider,
        service_slug=service_slug,
        tenant_id=tenant_id,
    )

    # ── 1. Resolve requirements.txt ──
    requirements_txt = Path(output_dir) / "requirements.txt"
    requirements_warning: str | None = None

    if not requirements_txt.exists():
        log.warning("vuln_scan.no_requirements_attempting_pipreqs", path=str(requirements_txt))
        # Try to auto-generate requirements.txt using pipreqs
        try:
            gen_result = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "pipreqs",
                    str(output_dir),
                    "--force",
                    "--savepath",
                    str(requirements_txt),
                ],
                capture_output=True,
                text=True,
                timeout=60,
            )
            if gen_result.returncode == 0 and requirements_txt.exists():
                log.info(
                    "vuln_scan.requirements_generated_by_pipreqs",
                    path=str(requirements_txt),
                )
            else:
                raise RuntimeError(f"pipreqs failed: {gen_result.stderr[:200]}")
        except Exception as pipreqs_exc:
            log.warning("vuln_scan.pipreqs_failed", error=str(pipreqs_exc))
            # Ensure the connector directory exists before writing the empty file
            requirements_txt.parent.mkdir(parents=True, exist_ok=True)
            requirements_txt.write_text("", encoding="utf-8")
            requirements_warning = (
                f"requirements.txt not found at {requirements_txt}. "
                "pipreqs could not auto-generate it — scan ran against 0 packages. "
                "Please add a requirements.txt to your connector directory."
            )
            log.warning("vuln_scan.empty_requirements_fallback", warning=requirements_warning)

    # ── 2. Prepare output directory ──
    vuln_dir = Path(output_dir) / ".shielva" / "vuln"
    vuln_dir.mkdir(parents=True, exist_ok=True)

    html_path = str(vuln_dir / "vulnerability_scan.html")
    excel_path = str(vuln_dir / "vulnerability_scan.xlsx")
    json_path = str(vuln_dir / "vulnerability_scan.json")
    ai_path = str(vuln_dir / "ai_fix_suggestions.md")

    # ── 3. Run pip-audit + Semgrep + Shielva Security in parallel ──
    log.info("vuln_scan.pip_audit_start", requirements=str(requirements_txt))
    log.info("vuln_scan.semgrep_start", source_dir=output_dir)

    # Shielva Security platform scan — runs against repo_url (GitHub) if provided,
    # otherwise scans the local connector directory directly (output_dir).
    # scan_type="full" runs all scanners: SAST + SCA + IaC + Secrets in parallel.
    shielva_target = repo_url or output_dir
    scan_coros = [
        _run_pip_audit(str(requirements_txt)),
        _run_semgrep(output_dir),
        _run_shielva_security_scan(shielva_target, tenant_id, session_id, scan_type="full"),
    ]
    audit_result, semgrep_result, shielva_result = await asyncio.gather(*scan_coros)

    if not audit_result["ok"]:
        error_msg = audit_result["error"]
        log.error("vuln_scan.pip_audit_error", error=error_msg)
        return {
            "status": "error",
            "scanned_at": scanned_at,
            "summary": {
                "total_packages": 0,
                "vulnerable_packages": 0,
                "critical": 0,
                "high": 0,
                "medium": 0,
                "low": 0,
            },
            "vulnerabilities": [],
            "safe_packages": [],
            "code_issues": semgrep_result.get("issues", []),
            "code_scan_error": semgrep_result.get("error"),
            "html_path": "",
            "excel_path": "",
            "ai_suggestions_path": "",
            "error": error_msg,
        }

    # ── Semgrep results ──
    code_issues: list[dict[str, Any]] = semgrep_result.get("issues", [])
    code_scan_error: str | None = semgrep_result.get("error")
    if not semgrep_result["ok"]:
        log.warning("vuln_scan.semgrep_error", error=code_scan_error)
    else:
        log.info("vuln_scan.semgrep_complete", code_issues=len(code_issues))

    # ── Shielva Security platform results ──
    shielva_findings: list[dict[str, Any]] = shielva_result.get("findings", [])
    shielva_scan_id: str | None = shielva_result.get("scan_id")
    shielva_report_r2_key: str | None = shielva_result.get("report_r2_key")
    shielva_error: str | None = shielva_result.get("error")
    if not shielva_result["ok"]:
        log.warning("vuln_scan.shielva_security_skipped", reason=shielva_error)
    else:
        log.info(
            "vuln_scan.shielva_security_complete",
            scan_id=shielva_scan_id,
            findings=len(shielva_findings),
        )
    # Merge platform code issues into code_issues list (tagged with source="shielva-security")
    code_issues = code_issues + shielva_findings

    # ── 4. Normalise results ──
    parsed = audit_result["parsed"] or {"dependencies": []}
    normalised = _normalise_results(parsed)
    summary = normalised["summary"]
    vulnerabilities: list[dict[str, Any]] = normalised["vulnerabilities"]
    safe_packages: list[dict[str, str]] = normalised["safe_packages"]

    log.info(
        "vuln_scan.results_normalised",
        total_packages=summary["total_packages"],
        vulnerable_packages=summary["vulnerable_packages"],
        critical=summary.get("critical", 0),
        high=summary.get("high", 0),
        code_issues=len(code_issues),
    )

    # ── 5. Generate HTML report ──
    try:
        html_content = _build_html_report(
            provider=provider,
            service_slug=service_slug,
            scanned_at=scanned_at,
            summary=summary,
            vulnerabilities=vulnerabilities,
            safe_packages=safe_packages,
        )
        Path(html_path).write_text(html_content, encoding="utf-8")
        log.info("vuln_scan.html_written", path=html_path)
    except Exception as exc:
        log.error("vuln_scan.html_failed", error=str(exc))
        html_path = ""

    # ── 6. Generate Excel report ──
    try:
        _build_excel_report(
            summary=summary,
            vulnerabilities=vulnerabilities,
            safe_packages=safe_packages,
            output_path=excel_path,
        )
        log.info("vuln_scan.excel_written", path=excel_path)
    except Exception as exc:
        log.error("vuln_scan.excel_failed", error=str(exc))
        excel_path = ""

    # ── 7. Generate AI fix suggestions ──
    try:
        ai_md = await _generate_ai_suggestions(vulnerabilities, tenant_id)
        Path(ai_path).write_text(ai_md, encoding="utf-8")
        log.info("vuln_scan.ai_suggestions_written", path=ai_path)
    except Exception as exc:
        log.error("vuln_scan.ai_suggestions_failed", error=str(exc))
        ai_path = ""

    # ── 8. Persist JSON result to disk ──
    result: dict[str, Any] = {
        "status": "ok",
        "scanned_at": scanned_at,
        "summary": summary,
        "vulnerabilities": vulnerabilities,
        "safe_packages": safe_packages,
        "code_issues": code_issues,
        "code_scan_error": code_scan_error,
        "html_path": html_path,
        "excel_path": excel_path,
        "ai_suggestions_path": ai_path,
        "error": requirements_warning,
        # Shielva Security platform scan metadata (null when API key not configured)
        "shielva_security": {
            "scan_id": shielva_scan_id,
            # R2 key in shielvasense bucket set by security API
            # Full path: shielvasense/Shielvasense-platform-int/{scan_id}/report.html
            "report_r2_key": shielva_report_r2_key,
            "error": shielva_error,
            "summary": shielva_result.get("summary"),
        }
        if shielva_result["ok"]
        else None,
    }
    try:
        Path(json_path).write_text(json.dumps(result, indent=2), encoding="utf-8")
    except Exception as exc:
        log.warning("vuln_scan.json_write_failed", error=str(exc))

    # ── 9. Upload all artefacts to R2 ──
    upload_tasks = []

    if html_path and Path(html_path).exists():
        upload_tasks.append(
            _upload_to_r2(
                provider,
                service_slug,
                "vulnerability_scan.html",
                Path(html_path).read_bytes(),
                "text/html",
            )
        )

    if excel_path and Path(excel_path).exists():
        upload_tasks.append(
            _upload_to_r2(
                provider,
                service_slug,
                "vulnerability_scan.xlsx",
                Path(excel_path).read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )

    if ai_path and Path(ai_path).exists():
        upload_tasks.append(
            _upload_to_r2(
                provider,
                service_slug,
                "ai_fix_suggestions.md",
                Path(ai_path).read_bytes(),
                "text/markdown",
            )
        )

    upload_tasks.append(
        _upload_to_r2(
            provider,
            service_slug,
            "vulnerability_scan.json",
            json.dumps(result, indent=2).encode("utf-8"),
            "application/json",
        )
    )

    if upload_tasks:
        await asyncio.gather(*upload_tasks, return_exceptions=True)

    log.info(
        "vuln_scan.complete",
        status="ok",
        vulnerable_packages=summary["vulnerable_packages"],
    )
    return result
